from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    CallbackContext,
    ConversationHandler,
    CallbackQueryHandler,
    JobQueue,
)
from telegram.constants import ChatMemberStatus
import re
import random
import openpyxl
from datetime import datetime, timedelta
import os

# Состояния
WAITING_FOR_START, FIND_IP, FIND_CITY, FIND_PASSWORD, FIND_HIDDEN_BUTTON, FINISH_QUEST = range(6)

# Данные для заданий
IP_ADDRESS = "95.52.96.86"  # Пример IP-адреса
CITY_NAME = "Кёнигсберг"  # Пример города
PASSWORD = "innohacker"  # Пример пароля

# Список случайных картинок для задания 3
RANDOM_IMAGES = [
    "https://usefoyer.com/ap/api/captcha?text=innohacker&type=text",
]

# Название файла Excel
EXCEL_FILE = "users_data.xlsx"

# ID администраторов, которые могут использовать команду /info
ADMIN_IDS = {5790141925, 634030933}

# Создаём файл Excel, если он не существует
def initialize_excel():
    try:
        # Пытаемся открыть существующий файл
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        # Если файл не существует, создаём новый
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        # Добавляем заголовки
        sheet.append(["ID", "Username", "Last Activity", "Current State", "Subscription Status"])
        # Сохраняем файл
        workbook.save(EXCEL_FILE)
        print(f"Файл создан по пути: {os.path.abspath(EXCEL_FILE)}")

# Сохраняем данные пользователя в Excel
async def save_user_data(user_id: int, username: str, current_state: str, subscription_status: str = None):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Проверяем, есть ли пользователь в базе
    user_found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_found = True
            break

    # Если пользователь уже есть, обновляем его данные
    if user_found:
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == user_id:
                row[1].value = username
                row[2].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                row[3].value = current_state
                if subscription_status:
                    row[4].value = subscription_status
                break
    else:
        # Если пользователя нет, добавляем новую запись
        sheet.append([user_id, username, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), current_state, subscription_status or "Не проверено"])

    # Сохраняем файл
    workbook.save(EXCEL_FILE)

# Проверяем, подписан ли пользователь на канал
async def check_subscription(user_id: int, context: CallbackContext) -> bool:
    try:
        chat_member = await context.bot.get_chat_member(chat_id="@innoprog", user_id=user_id)
        return chat_member.status in [ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR]
    except Exception as e:
        print(f"Ошибка при проверке подписки: {e}")
        return False

# Функция для старта бота
async def start(update: Update, context: CallbackContext) -> int:
    # Очищаем данные пользователя
    context.user_data.clear()

    # Инициализируем флаги заданий
    context.user_data["tasks"] = {
        "task1_completed": False,
        "task2_completed": False,
        "task3_completed": False,
    }

    # Сохраняем данные пользователя
    user_id = update.message.from_user.id
    username = update.message.from_user.username
    await save_user_data(user_id, username, "Старт квеста")

    # Отправляем описание квеста и кнопку "Я готов"
    reply_keyboard = [['Я готов!']]
    await update.message.reply_text(
        "**В одной крупной IT-компании случилась утечка данных.**\n\n"
        "Кто-то спрятал зашифрованный файл, и только умный хакер сможет его найти и разгадать! "
        "Тебе нужно пройти серию испытаний, чтобы спасти важную информацию.\n\n"
        "Готов ли ты начать?",
        reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True, resize_keyboard=True),
        parse_mode="Markdown"
    )

    # Запускаем напоминания
    if context.job_queue:
        context.job_queue.run_once(send_reminder, when=3600, data=user_id)  # Через 1 час
        context.job_queue.run_once(send_reminder, when=86400, data=user_id)  # Через 1 день
        context.job_queue.run_once(send_reminder, when=259200, data=user_id)  # Через 3 дня
    else:
        print("JobQueue не инициализирован. Убедитесь, что установлен пакет 'python-telegram-bot[job-queue]'.")

    return WAITING_FOR_START

# Функция для отправки напоминаний
async def send_reminder(context: CallbackContext):
    user_id = context.job.data
    await context.bot.send_message(
        chat_id=user_id,
        text="⏰ Ты всё ещё можешь продолжить квест и принять участие в розыгрыше! Не упусти свой шанс! 🚀"
    )

# Функция для обработки нажатия кнопки "Я готов!"
async def handle_start(update: Update, context: CallbackContext) -> int:
    # Удаляем предыдущие сообщения
    try:
        await context.bot.delete_message(chat_id=update.message.chat_id, message_id=update.message.message_id - 1)
        await context.bot.delete_message(chat_id=update.message.chat_id, message_id=update.message.message_id)
    except Exception as e:
        print(f"Ошибка при удалении сообщений: {e}")

    # Отправляем историю перед заданием 1
    await update.message.reply_text(
        "**История:**\n\n"
        "Ты получил(а) анонимное сообщение: 'В нашей компании произошла утечка данных. "
        "Кто-то спрятал зашифрованный файл на одном из наших серверов. "
        "Найди IP-адрес сервера, чтобы начать расследование.'\n\n"
        "Твоя задача — найти IP-адрес, спрятанный на сайте.",
        parse_mode="Markdown"
    )

    # Переходим к заданию с IP
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Перейти к заданию 1", url="https://bluvvis.github.io/it-quest/find_ip.html")]
    ])
    task_message = await update.message.reply_text(
        "**Задание 1:** Где-то на этом сайте спрятано число. Это IP-адрес. Найди его, и ты сделаешь первый шаг. 🔍",
        reply_markup=keyboard,
        parse_mode="Markdown"
    )

    # Сохраняем ID сообщения с заданием
    context.user_data["task_message_id"] = task_message.message_id
    return FIND_IP

# Функция для обработки задания 1 (поиск IP-адреса)
async def find_ip(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text

    # Проверка, что введённый текст — это IP-адрес
    if re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", user_input):
        if user_input == IP_ADDRESS:
            # Удаляем все сообщения, связанные с заданием
            await _delete_task_messages(update, context)

            # Устанавливаем флаг выполнения задания 1
            context.user_data["tasks"]["task1_completed"] = True

            # Сохраняем данные пользователя
            user_id = update.message.from_user.id
            username = update.message.from_user.username
            await save_user_data(user_id, username, "Задание 1 выполнено")

            # Отправляем историю перед заданием 2
            await update.message.reply_text(
                "**История:**\n\n"
                "Ты нашёл IP-адрес сервера. Теперь нужно определить, где он находится. ",
                parse_mode="Markdown"
            )

            # Отправляем сообщение с текстом задания 2
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("Перейти к заданию 2", url="https://www.google.com")]
            ])
            task_message = await update.message.reply_text(
                "**Задание 2:** Введи этот IP-адрес в Google и найди город, к которому он привязан. Название города в XIX веке — ключ ко следующему этапу! 🌍",
                reply_markup=keyboard,
                parse_mode="Markdown"
            )

            # Сохраняем ID нового сообщения с заданием
            context.user_data["task_message_id"] = task_message.message_id
            return FIND_CITY
        else:
            # Сохраняем ID сообщения с ошибкой
            error_message = await update.message.reply_text("Неправильный IP-адрес. Попробуй ещё раз! ❌")
            context.user_data.setdefault("error_message_ids", []).append(error_message.message_id)
            return FIND_IP
    else:
        # Сохраняем ID сообщения с ошибкой
        error_message = await update.message.reply_text("Это не похоже на IP-адрес. Попробуй ещё раз! ❌")
        context.user_data.setdefault("error_message_ids", []).append(error_message.message_id)
        return FIND_IP

# Функция для обработки задания 2 (поиск города)
async def find_city(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text

    # Приводим введённый текст к нижнему регистру и заменяем "ё" на "е"
    normalized_input = user_input.lower().replace("ё", "е")
    normalized_city_name = CITY_NAME.lower().replace("ё", "е")

    if normalized_input == normalized_city_name:
        # Удаляем все сообщения, связанные с заданием
        await _delete_task_messages(update, context)

        # Устанавливаем флаг выполнения задания 2
        context.user_data["tasks"]["task2_completed"] = True

        # Сохраняем данные пользователя
        user_id = update.message.from_user.id
        username = update.message.from_user.username
        await save_user_data(user_id, username, "Задание 2 выполнено")

        # Отправляем историю перед заданием 3
        await update.message.reply_text(
            "**История:**\n\n"
            "Ты определил город, где находится сервер. Теперь нужно получить доступ к системе. "
            ,
            parse_mode="Markdown"
        )

        # Отправляем сообщение с текстом задания 3
        task_message = await update.message.reply_text(
            "**Задание 3:** Мы сгенерировали captcha, попробуй расшифровать текст! 🔐",
            parse_mode="Markdown"
        )

        # Отправляем случайную картинку
        random_image = random.choice(RANDOM_IMAGES)
        photo_message = await update.message.reply_photo(random_image)

        # Сохраняем ID сообщения с заданием и картинки
        context.user_data["task_message_id"] = task_message.message_id
        context.user_data["photo_message_id"] = photo_message.message_id
        return FIND_PASSWORD
    else:
        # Сохраняем ID сообщения с ошибкой
        error_message = await update.message.reply_text("Неправильный город. Попробуй ещё раз! ❌")
        context.user_data.setdefault("error_message_ids", []).append(error_message.message_id)
        return FIND_CITY

# Функция для обработки задания 3 (поиск пароля)
async def find_password(update: Update, context: CallbackContext) -> int:
    user_input = update.message.text

    # Сравниваем пароль без учёта регистра
    if user_input.lower() == PASSWORD.lower():
        # Удаляем все сообщения, связанные с заданием
        await _delete_task_messages(update, context)

        # Устанавливаем флаг выполнения задания 3
        context.user_data["tasks"]["task3_completed"] = True

        # Сохраняем данные пользователя
        user_id = update.message.from_user.id
        username = update.message.from_user.username
        await save_user_data(user_id, username, "Задание 3 выполнено")

        # Отправляем историю перед заданием 4
        await update.message.reply_text(
            "**История:**\n\n"
            "Ты получил доступ к системе! Но файл защищён дополнительным уровнем безопасности. "
            ,
            parse_mode="Markdown"
        )

        # Отправляем сообщение с текстом задания 4
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("Перейти к заданию 4", url="https://bluvvis.github.io/it-quest/hidden_button.html")]
        ])
        await update.message.reply_text(
            "**Задание 4:** На сайте с паролем нужно найти «невидимую» кнопку и нажать на неё. 🕵️‍♂️",
            reply_markup=keyboard,
            parse_mode="Markdown"
        )
        return FIND_HIDDEN_BUTTON
    else:
        # Сохраняем ID сообщения с ошибкой
        error_message = await update.message.reply_text("Неправильный пароль. Попробуй ещё раз! ❌")
        context.user_data.setdefault("error_message_ids", []).append(error_message.message_id)
        return FIND_PASSWORD

# Функция для обработки задания 4 (поиск скрытой кнопки)
async def find_hidden_button(update: Update, context: CallbackContext) -> int:
    # Проверяем, что все задания 1-3 выполнены
    tasks = context.user_data.get("tasks", {})
    if not all(tasks.values()):
        await update.message.reply_text(
            "Ты ещё не прошёл все задания! Вернись и выполни их, чтобы завершить квест. ⚠️",
            parse_mode="Markdown"
        )
        return FIND_HIDDEN_BUTTON

    # Если все задания выполнены, завершаем квест
    await finish_quest(update, context)
    return ConversationHandler.END

# Функция для завершения квеста
async def finish_quest(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id
    username = update.message.from_user.username

    # Проверяем подписку сразу после завершения квеста
    is_subscribed = await check_subscription(user_id, context)
    subscription_status = "Подписан" if is_subscribed else "Не подписан"

    # Сохраняем данные пользователя в базу данных
    await save_user_data(user_id, username, "Квест завершён", subscription_status)

    # Отправляем поздравление и кнопки
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Подписаться на канал", url="https://t.me/innoprog")],
        [InlineKeyboardButton("Проверить подписку", callback_data="check_subscription")]
    ])
    await update.message.reply_text(
        "**Поздравляем! Ты успешно прошёл квест! 🎉**\n\n"
        "Для участия в розыгрыше и закрепления результата необходимо подписаться на наш канал:",
        reply_markup=keyboard,
        parse_mode="Markdown"
    )

# Функция для обработки проверки подписки
async def check_subscription_button(update: Update, context: CallbackContext):
    user_id = update.callback_query.from_user.id
    username = update.callback_query.from_user.username

    # Проверяем подписку
    is_subscribed = await check_subscription(user_id, context)

    if is_subscribed:
        # Обновляем статус подписки в Excel
        await save_user_data(user_id, username, "Подписка подтверждена", "Подписан")
        await update.callback_query.answer("Спасибо! Вы в списке участников. 🎉")
    else:
        # Обновляем статус подписки в Excel
        await save_user_data(user_id, username, "Подписка не подтверждена", "Не подписан")
        await update.callback_query.answer("Вы ещё не подписались на канал. ❌")

# Функция для отмены
async def cancel(update: Update, context: CallbackContext) -> int:
    await update.message.reply_text("Квест завершён. Если хочешь начать заново, введи /start.")
    return ConversationHandler.END

# Вспомогательная функция для удаления всех сообщений, связанных с заданием
async def _delete_task_messages(update: Update, context: CallbackContext):
    # Удаляем сообщение с заданием
    if "task_message_id" in context.user_data:
        await context.bot.delete_message(chat_id=update.message.chat_id, message_id=context.user_data["task_message_id"])
        del context.user_data["task_message_id"]

    # Удаляем сообщение с картинкой (если есть)
    if "photo_message_id" in context.user_data:
        await context.bot.delete_message(chat_id=update.message.chat_id, message_id=context.user_data["photo_message_id"])
        del context.user_data["photo_message_id"]

    # Удаляем все сообщения с ошибками
    if "error_message_ids" in context.user_data:
        for message_id in context.user_data["error_message_ids"]:
            try:
                await context.bot.delete_message(chat_id=update.message.chat_id, message_id=message_id)
            except Exception as e:
                print(f"Не удалось удалить сообщение: {e}")
        del context.user_data["error_message_ids"]

    # Удаляем последний ответ пользователя
    await context.bot.delete_message(chat_id=update.message.chat_id, message_id=update.message.message_id)

# Команда /info для выгрузки базы данных
async def info(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id
    if user_id not in ADMIN_IDS:
        await update.message.reply_text("У вас нет прав для выполнения этой команды. ❌")
        return

    # Отправляем файл Excel
    await update.message.reply_document(document=open(EXCEL_FILE, "rb"))

# Основная функция
def main() -> None:
    # Инициализируем файл Excel
    initialize_excel()

    # Вставь сюда свой токен
    application = Application.builder().token("8186767146:AAGIwAU1wulZVPQ8gN9w8YtMqUHJ8wq68lg").build()

    # Обработчик диалога
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            WAITING_FOR_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_start)],
            FIND_IP: [MessageHandler(filters.TEXT & ~filters.COMMAND, find_ip)],
            FIND_CITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, find_city)],
            FIND_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, find_password)],
            FIND_HIDDEN_BUTTON: [
                CommandHandler("finish", finish_quest),  # Обработчик команды /finish
                MessageHandler(filters.TEXT & ~filters.COMMAND, find_hidden_button),  # Обработчик текстовых сообщений
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    # Добавляем обработчик для кнопки проверки подписки
    application.add_handler(CallbackQueryHandler(check_subscription_button, pattern="^check_subscription$"))

    # Добавляем обработчик команды /info
    application.add_handler(CommandHandler("info", info))

    application.add_handler(conv_handler)

    # Запуск бота
    application.run_polling()

if __name__ == "__main__":
    main()
